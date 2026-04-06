import csv
import json
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

import build_reviewed_workbook_glossary as script


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "UI文本-已校对"
    ws.append(["string_id", "Chinese", "校对后"])
    ws.append(["1", "木叶入口", "«Коноха вход»"])
    ws.append(["2", "空白跳过", ""])
    ws.append(["3", "冲突词", "方案A"])
    ws.append(["4", "冲突词", "方案B"])
    ws.append(["5", "123", "Число"])
    ws.append(["6", "长长长长长长长长长长长长长长长长长长长", "太长"])
    ws.append(["7", "句子。", "Предложение"])
    ws.append(["8", "短词", "短  词  俄文"])

    ws2 = wb.create_sheet("system_toast-已校对")
    ws2.append(["string_id", "source_zh", "校对后"])
    ws2.append(["9", "木叶入口", "“Коноха вход”"])
    ws2.append(["10", "长度风险", "Очень длинный перевод"])

    ws3 = wb.create_sheet("Dialog-校对一部分")
    ws3.append(["string_id", "source_zh", "Ru"])
    ws3.append(["11", "不应纳入", "Не брать"])
    wb.save(path)


def _write_target_csv(path: Path) -> None:
    rows = [
        {"string_id": "1", "source_zh": "木叶入口", "target_text": "Старая Коноха вход", "target_ru": "Старая Коноха вход", "max_len_target": "20", "max_len_review_limit": "20"},
        {"string_id": "2", "source_zh": "短词", "target_text": "旧译", "target_ru": "旧译", "max_len_target": "6", "max_len_review_limit": "20"},
        {"string_id": "3", "source_zh": "长度风险", "target_text": "旧译", "target_ru": "旧译", "max_len_target": "", "max_len_review_limit": "8"},
        {"string_id": "4", "source_zh": "123", "target_text": "旧译", "target_ru": "旧译", "max_len_target": "10", "max_len_review_limit": "10"},
        {"string_id": "5", "source_zh": "句子。", "target_text": "旧译", "target_ru": "旧译", "max_len_target": "10", "max_len_review_limit": "10"},
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def test_sheet_filter_and_header_alias_and_blank_filter(tmp_path):
    workbook = tmp_path / "reviewed.xlsx"
    target_csv = tmp_path / "target.csv"
    out_dir = tmp_path / "out"
    _make_workbook(workbook)
    _write_target_csv(target_csv)

    assert script.main(["--workbook", str(workbook), "--target-csv", str(target_csv), "--out-dir", str(out_dir)]) == 0

    stats = json.loads((out_dir / "stats.json").read_text(encoding="utf-8"))
    assert stats["reviewed_sheet_count"] == 2
    assert stats["reviewed_record_count"] == 9


def test_conflict_split_and_resolved_yaml(tmp_path):
    workbook = tmp_path / "reviewed.xlsx"
    target_csv = tmp_path / "target.csv"
    out_dir = tmp_path / "out"
    _make_workbook(workbook)
    _write_target_csv(target_csv)

    script.main(["--workbook", str(workbook), "--target-csv", str(target_csv), "--out-dir", str(out_dir)])
    full_resolved = yaml.safe_load((out_dir / "full_resolved.yaml").read_text(encoding="utf-8"))
    conflicts = json.loads((out_dir / "conflicts.json").read_text(encoding="utf-8"))

    resolved_map = {entry["term_zh"]: entry["term_ru"] for entry in full_resolved["entries"]}
    assert "冲突词" not in resolved_map
    assert resolved_map["木叶入口"] == '"Коноха вход"'
    assert len(conflicts["conflicts"]) == 1
    assert conflicts["conflicts"][0]["source_zh"] == "冲突词"


def test_focus_filter_excludes_numeric_long_and_sentence_sources(tmp_path):
    workbook = tmp_path / "reviewed.xlsx"
    target_csv = tmp_path / "target.csv"
    out_dir = tmp_path / "out"
    _make_workbook(workbook)
    _write_target_csv(target_csv)

    script.main(["--workbook", str(workbook), "--target-csv", str(target_csv), "--out-dir", str(out_dir)])
    focus = yaml.safe_load((out_dir / "ui_art_focus_resolved.yaml").read_text(encoding="utf-8"))
    focus_terms = {entry["term_zh"] for entry in focus["entries"]}

    assert "木叶入口" in focus_terms
    assert "短词" in focus_terms
    assert "123" not in focus_terms
    assert "句子。" not in focus_terms
    assert "长长长长长长长长长长长长长长长长长长长" not in focus_terms


def test_compactness_gate_writes_manual_compact_conflicts(tmp_path):
    workbook = tmp_path / "reviewed.xlsx"
    target_csv = tmp_path / "target.csv"
    out_dir = tmp_path / "out"
    _make_workbook(workbook)
    _write_target_csv(target_csv)

    script.main(["--workbook", str(workbook), "--target-csv", str(target_csv), "--out-dir", str(out_dir)])
    focus = yaml.safe_load((out_dir / "ui_art_focus_resolved.yaml").read_text(encoding="utf-8"))
    focus_terms = {entry["term_zh"] for entry in focus["entries"]}
    compact_rows = list(csv.DictReader((out_dir / "manual_compact_conflicts.csv").open("r", encoding="utf-8-sig", newline="")))

    assert "长度风险" not in focus_terms
    assert any(row["source_zh"] == "长度风险" and row["reason"] == "reviewed_target_exceeds_max_len_review_limit" for row in compact_rows)


def test_compactness_gate_blocks_on_max_len_target_before_review_limit(tmp_path):
    workbook = tmp_path / "reviewed.xlsx"
    target_csv = tmp_path / "target.csv"
    out_dir = tmp_path / "out"
    _make_workbook(workbook)
    rows = [
        {
            "string_id": "1",
            "source_zh": "木叶入口",
            "target_text": "Старая Коноха вход",
            "target_ru": "Старая Коноха вход",
            "max_len_target": "6",
            "max_len_review_limit": "20",
        }
    ]
    with target_csv.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    script.main(["--workbook", str(workbook), "--target-csv", str(target_csv), "--out-dir", str(out_dir)])
    focus = yaml.safe_load((out_dir / "ui_art_focus_resolved.yaml").read_text(encoding="utf-8"))
    focus_terms = {entry["term_zh"] for entry in focus["entries"]}
    compact_rows = list(csv.DictReader((out_dir / "manual_compact_conflicts.csv").open("r", encoding="utf-8-sig", newline="")))

    assert "木叶入口" not in focus_terms
    assert any(row["source_zh"] == "木叶入口" and row["reason"] == "reviewed_target_exceeds_max_len_target" for row in compact_rows)


def test_normalize_text_collapses_space_and_quote_wrappers():
    assert script.normalize_text("«  文本  »") == '"文本"'
    assert script.normalize_text("短  词  俄文") == "短 词 俄文"
