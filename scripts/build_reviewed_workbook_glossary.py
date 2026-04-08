#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build authoritative and UI-art-focused glossary artifacts from the reviewed
Naruto RU workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import yaml
from openpyxl import load_workbook


SHEET_MARKER = "已校对"
SOURCE_HEADERS = ("Chinese", "source_zh")
TARGET_HEADER = "校对后"
SENTENCE_PUNCT_RE = re.compile(r"[，。！？；：、,.!?;:]")
PURE_DIGITS_RE = re.compile(r"^[0-9]+$")
PURE_ZH_NUMERAL_RE = re.compile(r"^[零一二三四五六七八九十百千万亿两〇]+$")
WHITESPACE_RE = re.compile(r"\s+")
WRAPPED_QUOTES = [
    ("«", "»"),
    ("“", "”"),
    ("‘", "’"),
    ('"', '"'),
    ("'", "'"),
]


@dataclass(frozen=True)
class ReviewedRecord:
    sheet_name: str
    row_number: int
    string_id: str
    source_zh: str
    target_ru: str


def configure_streams() -> None:
    if sys.platform != "win32" or "PYTEST_CURRENT_TEST" in os.environ:
        return
    import io

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\u00A0", " ").replace("\u200B", "")
    text = WHITESPACE_RE.sub(" ", text).strip()
    for left, right in WRAPPED_QUOTES:
        if text.startswith(left) and text.endswith(right) and len(text) >= 2:
            inner = text[len(left) : len(text) - len(right)].strip()
            text = f"\"{inner}\""
            break
    return WHITESPACE_RE.sub(" ", text).strip()


def source_is_focus_eligible(source_zh: str) -> bool:
    if not source_zh or len(source_zh) > 18:
        return False
    if SENTENCE_PUNCT_RE.search(source_zh):
        return False
    if PURE_DIGITS_RE.fullmatch(source_zh):
        return False
    if PURE_ZH_NUMERAL_RE.fullmatch(source_zh):
        return False
    return True


def load_reviewed_records(workbook_path: Path) -> Tuple[List[ReviewedRecord], List[Dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    records: List[ReviewedRecord] = []
    sheet_stats: List[Dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if SHEET_MARKER not in sheet_name:
            continue
        ws = workbook[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_text(value) for value in header_row]
        header_map = {header: index for index, header in enumerate(headers) if header}
        source_index = next((header_map.get(candidate) for candidate in SOURCE_HEADERS if candidate in header_map), None)
        target_index = header_map.get(TARGET_HEADER)
        string_id_index = header_map.get("string_id")
        if source_index is None or target_index is None:
            continue

        total_rows = 0
        usable_rows = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            source_zh = normalize_text(row[source_index] if source_index < len(row) else "")
            target_ru = normalize_text(row[target_index] if target_index < len(row) else "")
            if not source_zh or not target_ru:
                continue
            usable_rows += 1
            string_id = normalize_text(row[string_id_index] if string_id_index is not None and string_id_index < len(row) else "")
            records.append(
                ReviewedRecord(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    string_id=string_id,
                    source_zh=source_zh,
                    target_ru=target_ru,
                )
            )
        sheet_stats.append(
            {
                "sheet_name": sheet_name,
                "total_rows": total_rows,
                "usable_rows": usable_rows,
            }
        )
    return records, sheet_stats


def build_resolution_maps(records: Iterable[ReviewedRecord]) -> Tuple[Dict[str, Counter], Dict[str, List[ReviewedRecord]]]:
    target_counter: Dict[str, Counter] = defaultdict(Counter)
    provenance: Dict[str, List[ReviewedRecord]] = defaultdict(list)
    for record in records:
        target_counter[record.source_zh][record.target_ru] += 1
        provenance[record.source_zh].append(record)
    return dict(target_counter), dict(provenance)


def load_target_rows(target_csv: Path) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    with target_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            source_zh = normalize_text(row.get("source_zh") or "")
            if source_zh:
                grouped[source_zh].append(row)
    return dict(grouped)


def safe_int(value: Any) -> Optional[int]:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def build_resolved_entries(
    target_counter: Dict[str, Counter],
    provenance: Dict[str, List[ReviewedRecord]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    resolved: List[Dict[str, Any]] = []
    conflicts: List[Dict[str, Any]] = []
    for source_zh in sorted(target_counter.keys()):
        counter = target_counter[source_zh]
        rows = provenance[source_zh]
        sheets = sorted({row.sheet_name for row in rows})
        if len(counter) == 1:
            target_ru, support = counter.most_common(1)[0]
            resolved.append(
                {
                    "term_zh": source_zh,
                    "term_ru": target_ru,
                    "status": "approved",
                    "note": f"reviewed_workbook support={support}",
                    "support": support,
                    "sheet_names": sheets,
                    "source_occurrences": len(rows),
                }
            )
            continue
        conflicts.append(
            {
                "source_zh": source_zh,
                "variant_count": len(counter),
                "source_occurrences": len(rows),
                "sheet_names": sheets,
                "candidates": [
                    {"target_ru": target_ru, "support": support}
                    for target_ru, support in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
                ],
            }
        )
    return resolved, conflicts


def apply_focus_filters(
    resolved_entries: Iterable[Dict[str, Any]],
    target_rows: Dict[str, List[Dict[str, str]]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    focus_entries: List[Dict[str, Any]] = []
    compact_conflicts: List[Dict[str, Any]] = []
    for entry in resolved_entries:
        source_zh = str(entry["term_zh"])
        if source_zh not in target_rows:
            continue
        if not source_is_focus_eligible(source_zh):
            continue
        reviewed_target = str(entry["term_ru"])
        reviewed_len = len(reviewed_target)
        matched_rows = target_rows[source_zh]
        blocked = False
        for row in matched_rows:
            max_len_target = safe_int(row.get("max_length_target") or row.get("max_len_target"))
            max_len_review_limit = safe_int(row.get("max_len_review_limit"))
            if max_len_target is not None and reviewed_len > max_len_target:
                blocked = True
                compact_conflicts.append(
                    {
                        "string_id": str(row.get("string_id") or "").strip(),
                        "source_zh": source_zh,
                        "reviewed_target_ru": reviewed_target,
                        "current_target": str(row.get("target_text") or row.get("target_ru") or row.get("target") or "").strip(),
                        "max_len_target": max_len_target,
                        "max_len_review_limit": max_len_review_limit,
                        "reviewed_target_len": reviewed_len,
                        "overflow": reviewed_len - max_len_target,
                        "reason": "reviewed_target_exceeds_max_len_target",
                    }
                )
                continue
            if max_len_review_limit is not None and reviewed_len > max_len_review_limit:
                blocked = True
                compact_conflicts.append(
                    {
                        "string_id": str(row.get("string_id") or "").strip(),
                        "source_zh": source_zh,
                        "reviewed_target_ru": reviewed_target,
                        "current_target": str(row.get("target_text") or row.get("target_ru") or row.get("target") or "").strip(),
                        "max_len_target": max_len_target,
                        "max_len_review_limit": max_len_review_limit,
                        "reviewed_target_len": reviewed_len,
                        "overflow": reviewed_len - max_len_review_limit,
                        "reason": "reviewed_target_exceeds_max_len_review_limit",
                    }
                )
        if blocked:
            continue
        focus_entries.append(
            {
                "term_zh": source_zh,
                "term_ru": reviewed_target,
                "status": "approved",
                "note": str(entry.get("note") or ""),
                "support": int(entry.get("support") or 0),
                "matched_row_count": len(matched_rows),
            }
        )
    return focus_entries, compact_conflicts


def write_csv(path: Path, rows: Iterable[Dict[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def dump_yaml(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        yaml.safe_dump(payload, fh, allow_unicode=True, sort_keys=False, default_flow_style=False)


def dump_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build reviewed workbook glossary artifacts")
    parser.add_argument("--workbook", required=True, help="Reviewed workbook path")
    parser.add_argument("--target-csv", required=True, help="Residual v2 base full CSV for UI-art focus filtering")
    parser.add_argument("--out-dir", required=True, help="Output directory for glossary artifacts")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    configure_streams()
    args = parse_args(argv)
    workbook_path = Path(args.workbook)
    target_csv = Path(args.target_csv)
    out_dir = Path(args.out_dir)

    records, sheet_stats = load_reviewed_records(workbook_path)
    target_counter, provenance = build_resolution_maps(records)
    resolved_entries, conflicts = build_resolved_entries(target_counter, provenance)
    target_rows = load_target_rows(target_csv)
    focus_entries, compact_conflicts = apply_focus_filters(resolved_entries, target_rows)

    dump_yaml(
        out_dir / "full_resolved.yaml",
        {
            "meta": {
                "type": "approved",
                "source": str(workbook_path),
                "count": len(resolved_entries),
            },
            "entries": resolved_entries,
        },
    )
    dump_yaml(
        out_dir / "ui_art_focus_resolved.yaml",
        {
            "meta": {
                "type": "approved",
                "source": str(workbook_path),
                "focus_target_csv": str(target_csv),
                "count": len(focus_entries),
            },
            "entries": focus_entries,
        },
    )
    dump_json(out_dir / "conflicts.json", {"conflicts": conflicts})
    write_csv(out_dir / "conflicts.csv", [
        {
            "source_zh": item["source_zh"],
            "target_ru": candidate["target_ru"],
            "support": candidate["support"],
            "variant_count": item["variant_count"],
            "source_occurrences": item["source_occurrences"],
            "sheet_names": "|".join(item["sheet_names"]),
        }
        for item in conflicts
        for candidate in item["candidates"]
    ], ["source_zh", "target_ru", "support", "variant_count", "source_occurrences", "sheet_names"])
    write_csv(
        out_dir / "manual_compact_conflicts.csv",
        compact_conflicts,
        [
            "string_id",
            "source_zh",
            "reviewed_target_ru",
            "current_target",
            "max_len_target",
            "max_len_review_limit",
            "reviewed_target_len",
            "overflow",
            "reason",
        ],
    )
    dump_json(
        out_dir / "stats.json",
        {
            "workbook": str(workbook_path),
            "target_csv": str(target_csv),
            "reviewed_sheet_count": len(sheet_stats),
            "sheet_stats": sheet_stats,
            "reviewed_record_count": len(records),
            "unique_source_count": len(target_counter),
            "resolved_source_count": len(resolved_entries),
            "conflict_source_count": len(conflicts),
            "focus_candidate_source_count": len([entry for entry in resolved_entries if entry["term_zh"] in target_rows]),
            "focus_resolved_count": len(focus_entries),
            "compact_conflicts": len(compact_conflicts),
        },
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
